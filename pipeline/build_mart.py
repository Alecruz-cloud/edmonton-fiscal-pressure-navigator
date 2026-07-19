"""Edmonton Fiscal Pressure Navigator — fixed extraction pipeline.

Deterministic raw -> curated build. Reads the raw workbooks READ-ONLY and writes:

    data/curated/financial_long.csv   all municipalities x approved codes x 2018-2025
    data/curated/population.csv       all CSD rows x 2016-2025
    data/curated/metric_mart.csv      year x municipality x metric (Edmonton, Calgary, AB-cities median)
    data/curated/metric_mart.json     same records + run metadata for the dashboard

Load-bearing rules (CLAUDE.md / governance/metric_dictionary.md):
  * extraction is by worksheet (schedule) + variable code found in row 3 — never column position;
  * municipality CODE is text (leading zeros preserved);
  * per-capita denominators come only from the population-estimates workbook;
  * blank stays `missing`, explicit 0 stays `reported_zero` — never merged;
  * dollar values are nominal; every value carries a full source trace.
"""
from __future__ import annotations

import csv
import io
import json
import statistics
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw"
CURATED = ROOT / "data" / "curated"
POP_FILE = "population-estimates-ab-census-subdivision-municipal-2016-to-current.xlsx"
# StatCan table 18-10-0005-01 (annual CPI, All-items, Alberta, 2002=100), downloaded
# 2026-07-18 — owner-approved the same day for DISPLAY-ONLY real-dollar context
# (base 2025). The signal engine evaluates nominal values under rules v1.2.
CPI_FILE = "18100005-eng.zip"
CPI_BASE_YEAR = 2025

YEARS = list(range(2018, 2026))

# Approved extraction spec — mirrors governance/metric_dictionary.md. Do not add codes here
# without owner approval of the dictionary first.
EXTRACT = {
    "C(1)-Revenue": ["01140", "01580", "01590", "01210", "01220", "01290", "01310", "01530", "01540"],
    "AA(1)-Debt": ["05700", "05710", "05720", "05730"],
    "A(1)-Total": ["00395", "00450"],
    "EA(1)-Assessment": ["08260"],
    "K(3)-Total": ["04000"],
    "MR(3)-Mill Rate": ["10030", "10034"],
    "ST(1)-Stat": ["05500"],
}

BASE_METRICS = {  # metric_id -> (sheet, code, name, domain, unit)
    "rev_total": ("C(1)-Revenue", "01140", "Total revenue", "A", "$ nominal"),
    "exp_total": ("C(1)-Revenue", "01580", "Total expense", "A", "$ nominal"),
    "net_rev_exp": ("C(1)-Revenue", "01590", "Net revenue/expense", "A", "$ nominal"),
    "debt_total": ("AA(1)-Debt", "05710", "Total debt", "B", "$ nominal"),
    "debt_limit": ("AA(1)-Debt", "05700", "Debt limit", "B", "$ nominal"),
    "debt_service_cost": ("AA(1)-Debt", "05730", "Total debt-service costs", "B", "$ nominal"),
    "debt_service_limit": ("AA(1)-Debt", "05720", "Debt-service limit", "B", "$ nominal"),
    "net_financial_assets": ("A(1)-Total", "00395", "Net financial assets / net debt", "B", "$ nominal"),
    "accumulated_surplus": ("A(1)-Total", "00450", "Accumulated surplus", "B", "$ nominal"),
    "equalized_assessment": ("EA(1)-Assessment", "08260", "Equalized assessment", "C", "$ nominal"),
    "property_tax": ("K(3)-Total", "04000", "Property taxes and grants in place", "C", "$ nominal"),
    "mill_rate_res": ("MR(3)-Mill Rate", "10030", "Residential general municipal mill rate", "C", "mills"),
    "mill_rate_nonres": ("MR(3)-Mill Rate", "10034", "Non-residential general municipal mill rate", "C", "mills"),
    "exp_police": ("C(1)-Revenue", "01210", "Police expense", "D", "$ nominal"),
    "exp_fire": ("C(1)-Revenue", "01220", "Fire expense", "D", "$ nominal"),
    "exp_roads": ("C(1)-Revenue", "01290", "Roads, streets, walks, lighting expense", "D", "$ nominal"),
    "exp_transit": ("C(1)-Revenue", "01310", "Public transit expense", "D", "$ nominal"),
    "exp_parks_rec": ("C(1)-Revenue", "01530", "Parks and recreation expense", "D", "$ nominal"),
    "exp_culture": ("C(1)-Revenue", "01540", "Culture expense", "D", "$ nominal"),
    "fte_total": ("ST(1)-Stat", "05500", "Total full-time municipal positions", "D", "positions"),
}

# Derived metrics: metric_id -> (numerator id, denominator id, name, domain, unit, formula, scale)
DERIVED = {
    "debt_utilization": ("debt_total", "debt_limit", "Debt utilization", "B", "%",
                         "debt_total / debt_limit", 100.0),
    "debt_service_utilization": ("debt_service_cost", "debt_service_limit", "Debt-service utilization", "B", "%",
                                 "debt_service_cost / debt_service_limit", 100.0),
    "rev_per_capita": ("rev_total", "pop", "Revenue per capita", "A", "$ nominal / person",
                       "rev_total / pop", 1.0),
    "exp_per_capita": ("exp_total", "pop", "Expense per capita", "A", "$ nominal / person",
                       "exp_total / pop", 1.0),
    "assessment_per_capita": ("equalized_assessment", "pop", "Assessment per capita", "C", "$ nominal / person",
                              "equalized_assessment / pop", 1.0),
    "exp_police_per_capita": ("exp_police", "pop", "Police expense per capita", "D", "$ nominal / person",
                              "exp_police / pop", 1.0),
    "exp_fire_per_capita": ("exp_fire", "pop", "Fire expense per capita", "D", "$ nominal / person",
                            "exp_fire / pop", 1.0),
    "exp_roads_per_capita": ("exp_roads", "pop", "Roads expense per capita", "D", "$ nominal / person",
                             "exp_roads / pop", 1.0),
    "exp_transit_per_capita": ("exp_transit", "pop", "Transit expense per capita", "D", "$ nominal / person",
                               "exp_transit / pop", 1.0),
    "exp_parks_rec_per_capita": ("exp_parks_rec", "pop", "Parks & recreation expense per capita", "D",
                                 "$ nominal / person", "exp_parks_rec / pop", 1.0),
    "exp_culture_per_capita": ("exp_culture", "pop", "Culture expense per capita", "D", "$ nominal / person",
                               "exp_culture / pop", 1.0),
    "exp_police_share": ("exp_police", "exp_total", "Police share of total expense", "D", "%",
                         "exp_police / exp_total", 100.0),
    "exp_fire_share": ("exp_fire", "exp_total", "Fire share of total expense", "D", "%",
                       "exp_fire / exp_total", 100.0),
    "exp_roads_share": ("exp_roads", "exp_total", "Roads share of total expense", "D", "%",
                        "exp_roads / exp_total", 100.0),
    "exp_transit_share": ("exp_transit", "exp_total", "Transit share of total expense", "D", "%",
                          "exp_transit / exp_total", 100.0),
    "exp_parks_rec_share": ("exp_parks_rec", "exp_total", "Parks & recreation share of total expense", "D", "%",
                            "exp_parks_rec / exp_total", 100.0),
    "exp_culture_share": ("exp_culture", "exp_total", "Culture share of total expense", "D", "%",
                          "exp_culture / exp_total", 100.0),
    "fte_per_1000": ("fte_total", "pop", "FTE per 1,000 residents", "D", "per 1,000 residents",
                     "fte_total / (pop / 1000)", 1000.0),
}

# Focus municipalities (governance/municipality_crosswalk.csv). These two are the only
# matches approved as population joins; the crosswalk's 17 other city rows are
# confirmed_geo_only (context-map placement, governance/crosswalk_signoff.md) and must
# never be used as per-capita denominators.
FOCUS = {"0098": ("Edmonton", "4811061"), "0046": ("Calgary", "4806016")}

# Metrics for which an Alberta reporting-city median is computable from the financial
# workbooks alone (no population join, which would need unapproved crosswalk matches).
MEDIAN_METRICS = ["debt_utilization", "debt_service_utilization", "mill_rate_res", "mill_rate_nonres"]


def parse_value(cell):
    """Classify a raw cell -> (numeric_or_None, quality_status)."""
    if cell is None or (isinstance(cell, str) and cell.strip() == ""):
        return None, "missing"
    if isinstance(cell, (int, float)):
        num = float(cell)
    else:
        try:
            num = float(str(cell).replace(",", "").strip())
        except ValueError:
            return None, "not_applicable"
    if num == 0:
        return 0.0, "reported_zero"
    return num, "reported_value"


def extract_financial():
    """All municipalities x approved codes x years -> list of long-form records."""
    records = []
    for year in YEARS:
        fname = f"{year}_financial_year.xlsx"
        wb = openpyxl.load_workbook(RAW / fname, read_only=True)
        for sheet, codes in EXTRACT.items():
            ws = wb[sheet]
            rows = ws.iter_rows(values_only=True)
            next(rows)                      # row 1: title
            labels = next(rows)             # row 2: variable labels
            code_row = next(rows)           # row 3: variable codes
            colmap = {}
            for idx, c in enumerate(code_row):
                s = str(c).strip() if c is not None else ""
                if s in codes:
                    colmap[s] = idx
            missing_codes = [c for c in codes if c not in colmap]
            if missing_codes:
                raise RuntimeError(f"{year} {sheet}: approved codes not found in row 3: {missing_codes}")
            for row in rows:                # data rows
                code = row[2] if len(row) > 2 else None
                if code is None or not str(code).strip():
                    continue
                mcode = str(code).strip()
                status = str(row[1]).strip() if row[1] is not None else ""
                name = str(row[3]).strip() if row[3] is not None else ""
                for vcode, idx in sorted(colmap.items()):
                    cell = row[idx] if idx < len(row) else None
                    num, quality = parse_value(cell)
                    records.append({
                        "financial_year": year,
                        "municipality_code": mcode,
                        "municipality_name_raw": name,
                        "municipality_status": status,
                        "schedule": sheet,
                        "variable_code": vcode,
                        "variable_name": str(labels[idx]).strip() if idx < len(labels) and labels[idx] else "",
                        "value": num,
                        "quality_status": quality,
                        "source_file": fname,
                        "source_sheet": sheet,
                    })
        wb.close()
    return records


def extract_population():
    """All CSD rows x 2016-2025 -> list of records."""
    records = []
    wb = openpyxl.load_workbook(RAW / POP_FILE, read_only=True)
    ws = wb["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    header = None
    for row in rows:
        if row[0] is not None and str(row[0]).strip() == "Census Division":
            header = [str(c).strip() if c is not None else "" for c in row]
            break
    if header is None:
        raise RuntimeError("Population header row not found")
    year_cols = {int(h): i for i, h in enumerate(header) if h.isdigit()}
    for row in rows:
        csd = str(row[1]).strip() if row[1] is not None else ""
        if not csd.isdigit():
            continue  # footnote / blank rows
        name_raw = str(row[2]).strip() if row[2] is not None else ""
        for year, idx in sorted(year_cols.items()):
            num, quality = parse_value(row[idx] if idx < len(row) else None)
            records.append({
                "year": year,
                "census_subdivision_code": csd,
                "municipality_name_raw": name_raw,
                "municipality_type": str(row[3]).strip() if row[3] is not None else "",
                "population": int(num) if num is not None else None,
                "quality_status": quality,
                "source_file": POP_FILE,
                "source_sheet": "Sheet1",
            })
    wb.close()
    return records


def extract_cpi():
    """Alberta All-items annual CPI (2002=100) -> {year: value}, read from the raw zip.

    StatCan table 18-10-0005-01, owner-approved 2026-07-18 for display-only
    real-dollar context. Hard-fails unless exactly one clean value exists for
    every project year — a status flag or unit change must stop the build.
    """
    cpi = {}
    with zipfile.ZipFile(RAW / CPI_FILE) as z, z.open("18100005.csv") as fh:
        for r in csv.DictReader(io.TextIOWrapper(fh, encoding="utf-8-sig")):
            if (r["GEO"] == "Alberta" and r["Products and product groups"] == "All-items"
                    and r["REF_DATE"].isdigit() and int(r["REF_DATE"]) in set(YEARS)):
                year = int(r["REF_DATE"])
                if r["UOM"] != "2002=100":
                    raise RuntimeError(f"CPI {year}: unexpected unit of measure {r['UOM']!r}")
                if r["STATUS"].strip():
                    raise RuntimeError(f"CPI {year}: unexpected status flag {r['STATUS']!r}")
                if year in cpi:
                    raise RuntimeError(f"CPI {year}: duplicate source row")
                cpi[year] = float(r["VALUE"])
    missing = [y for y in YEARS if y not in cpi]
    if missing:
        raise RuntimeError(f"{CPI_FILE}: CPI years missing for Alberta All-items: {missing}")
    return cpi


def build_mart(fin, pop, cpi):
    """Narrow decision-ready mart: Edmonton + Calgary + AB-cities median, plus
    Edmonton-only constant-2025$ display context (owner-approved 2026-07-18)."""
    fin_ix = {(r["financial_year"], r["municipality_code"], r["schedule"], r["variable_code"]): r
              for r in fin}
    pop_ix = {(r["year"], r["census_subdivision_code"]): r for r in pop}

    mart = []

    def add(year, mcode, mname, mid, name, domain, value, unit, schedule, vcode,
            quality, note, formula="", comparability=None):
        mart.append({
            "year": year, "municipality_code": mcode, "municipality_name": mname,
            "metric_id": mid, "metric_name": name, "metric_domain": domain,
            "value": value, "unit": unit,
            "source_file": (POP_FILE if schedule == "population" else
                            f"{year}_financial_year.xlsx" if schedule not in ("derived",) else "derived"),
            "source_schedule": schedule, "source_variable_code": vcode,
            "quality_status": quality,
            "comparability_status": comparability or (
                "provisional_peer_year" if (year == 2025 and mcode == "AB_CITIES_MEDIAN")
                else "comparable"),
            "formula": formula, "data_quality_note": note,
        })

    for year in YEARS:
        for mcode, (mname, csd) in sorted(FOCUS.items()):
            # population
            prow = pop_ix.get((year, csd))
            pop_val = prow["population"] if prow else None
            add(year, mcode, mname, "pop", "Population", "A", pop_val, "persons",
                "population", f"CSD {csd}",
                prow["quality_status"] if prow else "missing",
                "July 1 estimate; sole approved per-capita denominator")
            prev = pop_ix.get((year - 1, csd))
            if prow and prev and prow["population"] and prev["population"]:
                add(year, mcode, mname, "pop_growth", "Annual population growth", "A",
                    round((prow["population"] / prev["population"] - 1) * 100, 4), "%",
                    "derived", f"CSD {csd}", "reported_value",
                    "derived from population estimates", "pop[y] / pop[y-1] - 1")
            # base metrics
            base_vals = {"pop": float(pop_val) if pop_val else None}
            for mid, (sheet, vcode, name, domain, unit) in sorted(BASE_METRICS.items()):
                r = fin_ix.get((year, mcode, sheet, vcode))
                val = r["value"] if r else None
                q = r["quality_status"] if r else "missing"
                base_vals[mid] = val
                note = "nominal dollars" if unit.startswith("$") else ""
                add(year, mcode, mname, mid, name, domain, val, unit, sheet, vcode, q, note)
            # derived metrics
            derived_vals = {}
            for mid, (num_id, den_id, name, domain, unit, formula, scale) in sorted(DERIVED.items()):
                num, den = base_vals.get(num_id), base_vals.get(den_id)
                if num is None or den is None:
                    add(year, mcode, mname, mid, name, domain, None, unit, "derived",
                        f"{num_id}/{den_id}", "not_applicable",
                        "input missing — not computed (missing is never treated as zero)", formula)
                elif den == 0:
                    add(year, mcode, mname, mid, name, domain, None, unit, "derived",
                        f"{num_id}/{den_id}", "not_applicable", "denominator reported zero", formula)
                else:
                    derived_vals[mid] = round(num / den * scale, 6)
                    add(year, mcode, mname, mid, name, domain, derived_vals[mid],
                        unit, "derived", f"{num_id}/{den_id}", "reported_value",
                        "nominal inputs" if "$" in unit else "", formula)

            # Real-dollar DISPLAY context (owner-approved 2026-07-18): Edmonton only.
            # Published nominal value × CPI_base/CPI_year, base 2025. These rows exist
            # for the drill-down toggle — the signal engine evaluates nominal values
            # under rules v1.2 and never reads *_real2025 metrics.
            if mcode == "0098":
                factor = cpi[CPI_BASE_YEAR] / cpi[year]
                dollar_ids = ([m for m, s in sorted(BASE_METRICS.items()) if s[4].startswith("$")]
                              + [m for m, s in sorted(DERIVED.items()) if s[4].startswith("$")])
                for mid in dollar_ids:
                    spec = BASE_METRICS.get(mid) or DERIVED.get(mid)
                    name, domain, unit = spec[2], spec[3], spec[4]
                    real_unit = unit.replace("$ nominal", f"$ constant {CPI_BASE_YEAR}")
                    nominal = base_vals.get(mid) if mid in BASE_METRICS else derived_vals.get(mid)
                    if nominal is None:
                        add(year, mcode, mname, f"{mid}_real{CPI_BASE_YEAR}",
                            f"{name} (constant {CPI_BASE_YEAR} $)", domain, None, real_unit,
                            "derived", f"{mid} x CPI{CPI_BASE_YEAR}/CPI{year}", "not_applicable",
                            "nominal input not computed — not deflated (missing is never treated as zero)",
                            f"{mid} × (CPI_{CPI_BASE_YEAR} / CPI_year)",
                            comparability="display_only_context")
                    else:
                        add(year, mcode, mname, f"{mid}_real{CPI_BASE_YEAR}",
                            f"{name} (constant {CPI_BASE_YEAR} $)", domain,
                            round(nominal * factor, 6), real_unit,
                            "derived", f"{mid} x CPI{CPI_BASE_YEAR}/CPI{year}", "reported_value",
                            f"deflated with StatCan 18-10-0005-01, All-items Alberta annual CPI "
                            f"(2002=100: {cpi[year]} → {cpi[CPI_BASE_YEAR]}), base {CPI_BASE_YEAR} — "
                            "display-only context, owner-approved 2026-07-18; signals evaluate "
                            "nominal values under rules v1.2",
                            f"{mid} × (CPI_{CPI_BASE_YEAR} / CPI_year)",
                            comparability="display_only_context")

        # AB reporting-city median (financial-workbook-only metrics; no population joins).
        # rules v1.2: the subject municipality (Edmonton, 0098) is excluded from its own
        # comparator, and the included municipalities are named, not just counted.
        SUBJECT = "0098"
        city_rows = {(r["municipality_code"], r["municipality_name_raw"]) for r in fin
                     if r["financial_year"] == year and r["municipality_status"] == "City"}
        cities = sorted(c for c, _ in city_rows if c != SUBJECT)
        name_of = {c: n for c, n in city_rows}
        for mid in MEDIAN_METRICS:
            vals, included = [], []
            for c in cities:
                v = None
                if mid in DERIVED:
                    num_id, den_id = DERIVED[mid][0], DERIVED[mid][1]
                    ns, nc = BASE_METRICS[num_id][0], BASE_METRICS[num_id][1]
                    ds, dc = BASE_METRICS[den_id][0], BASE_METRICS[den_id][1]
                    nr, dr = fin_ix.get((year, c, ns, nc)), fin_ix.get((year, c, ds, dc))
                    if nr and dr and nr["value"] is not None and dr["value"] not in (None, 0.0):
                        v = nr["value"] / dr["value"] * DERIVED[mid][6]
                else:
                    sheet, vcode = BASE_METRICS[mid][0], BASE_METRICS[mid][1]
                    r = fin_ix.get((year, c, sheet, vcode))
                    if r and r["quality_status"] == "reported_value":
                        v = r["value"]
                if v is not None:
                    vals.append(v)
                    included.append(name_of[c].title())
            name = DERIVED[mid][2] if mid in DERIVED else BASE_METRICS[mid][2]
            domain = DERIVED[mid][3] if mid in DERIVED else BASE_METRICS[mid][3]
            unit = DERIVED[mid][4] if mid in DERIVED else BASE_METRICS[mid][4]
            add(year, "AB_CITIES_MEDIAN",
                f"Alberta reporting cities excl. Edmonton (n={len(vals)})",
                mid, name, domain,
                round(statistics.median(vals), 6) if vals else None, unit,
                "derived", "median across reporting cities excl. subject",
                "reported_value" if vals else "not_applicable",
                f"median of {len(vals)} reporting cities in {year}, subject (Edmonton) excluded "
                f"per rules v1.2; included: {', '.join(included) if included else 'none'}; "
                "supporting context only — cannot create a high-priority signal",
                f"median({mid})")

        # Per-capita medians are structurally unavailable: population joins are approved
        # for Edmonton and Calgary only. The crosswalk's other city rows are
        # confirmed_geo_only (context-map placement) and are not population joins.
        # Emit explicit rows so the state is visible, never silent.
        for mid, spec in sorted(DERIVED.items()):
            if spec[1] != "pop":
                continue
            add(year, "AB_CITIES_MEDIAN", "Alberta reporting cities",
                mid, spec[2], spec[3], None, spec[4],
                "derived", "median across reporting cities",
                "structurally_unavailable",
                "median not computable — population joins in the municipality crosswalk are "
                "approved for Edmonton and Calgary only; the other city matches are "
                "geo/context-map scope (confirmed_geo_only) and are never used as "
                "per-capita denominators",
                f"median({mid})")

    return mart


def coverage(fin):
    """Returns received per year (unique municipality codes in C(1)-Revenue)."""
    out = {}
    for year in YEARS:
        out[year] = len({r["municipality_code"] for r in fin
                         if r["financial_year"] == year and r["schedule"] == "C(1)-Revenue"})
    return out


def write_csv(path, records, fields):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(records)


def main():
    CURATED.mkdir(parents=True, exist_ok=True)
    fin = extract_financial()
    pop = extract_population()
    cpi = extract_cpi()
    mart = build_mart(fin, pop, cpi)
    cov = coverage(fin)

    write_csv(CURATED / "financial_long.csv", fin, list(fin[0].keys()))
    write_csv(CURATED / "population.csv", pop, list(pop[0].keys()))
    write_csv(CURATED / "metric_mart.csv", mart, list(mart[0].keys()))

    meta = {
        "built_from": "data/raw (read-only)",
        "governance": ["governance/metric_dictionary.md", "governance/municipality_crosswalk.csv"],
        "dollar_basis": "nominal — signals evaluate nominal values (rules v1.2); "
                        "constant-2025$ display context available on *_real2025 metrics",
        "population_source": POP_FILE,
        "returns_received_by_year": cov,
        "coverage_note_2025": f"{cov[2025]} of 332 financial information returns received — peer results "
                              "provisional (332 per the 2025 workbook's AA(1)-Debt footer row "
                              f"'{cov[2025]} out of 332')",
        "quality_states": ["reported_value", "reported_zero", "missing", "not_applicable",
                           "structurally_unavailable"],
        "cpi_context": {
            "source_file": CPI_FILE, "table": "18-10-0005-01",
            "series": "Consumer Price Index, annual average, All-items, Alberta "
                      "(2002=100), vector v41694625",
            "base_year": CPI_BASE_YEAR, "values": cpi,
            "scope": "display-only context, owner-approved 2026-07-18 — signals "
                     "evaluate nominal values (rules v1.2)",
        },
    }
    with open(CURATED / "metric_mart.json", "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "records": mart}, f, indent=1)

    print(f"financial_long: {len(fin):,} rows | population: {len(pop):,} rows | mart: {len(mart):,} rows")
    print("returns received by year:", cov)
    ed25 = {r["metric_id"]: r["value"] for r in mart
            if r["year"] == 2025 and r["municipality_code"] == "0098"}
    print("Edmonton 2025 spot-check:",
          {k: ed25.get(k) for k in ("pop", "debt_total", "debt_limit", "debt_utilization",
                                    "exp_total", "exp_per_capita", "fte_per_1000")})
    print("CPI Alberta (2002=100):", cpi, "| base year:", CPI_BASE_YEAR)
    real18 = next((r["value"] for r in mart if r["year"] == 2018
                   and r["municipality_code"] == "0098"
                   and r["metric_id"] == f"exp_per_capita_real{CPI_BASE_YEAR}"), None)
    print(f"real spot-check: exp_per_capita_real{CPI_BASE_YEAR} 2018 = {real18} "
          f"(2025 identity = {ed25.get(f'exp_per_capita_real{CPI_BASE_YEAR}')})")


if __name__ == "__main__":
    main()
